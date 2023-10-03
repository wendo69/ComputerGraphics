import pandas as pd
import openpyxl as xlsx
from pandas import DataFrame
import os



def variables():
    # This function returns an array of dataframes of all the data is the dataset folder of all the languages
    dataframes = []

    directory = 'dataset\data'
    for filename in os.scandir(directory):
        if filename.is_file():
            dataframes.append({'name': filename.name, 'data': pd.read_json(path_or_buf=filename.path, lines=True)})
    return dataframes

variables = variables()

def english_translations_to_xlsx(file_path, language_abbv):
    # This function returns an .xlsx file that has translations from english to the specified language

    # Import our languages, English(the pivot), and the specified language
    df_pivot = pd.read_json(path_or_buf='dataset/data/en-US.jsonl', lines=True, encoding='utf-8').get(
        ['id', 'utt', 'annot_utt'])
    df_lang = pd.read_json(path_or_buf=file_path, lines=True, encoding='utf-8').get(['id', 'utt', 'annot_utt'])

    # Create a workbook and set an active worksheet which will be used to store our data
    workbook = xlsx.Workbook()
    worksheet = workbook.active
    # Adding Headers to the Sheet
    
    worksheet['A1'] = 'id'
    worksheet['B1'] = 'utt'
    worksheet['C1'] = 'annot_utt'
    worksheet['D1'] = file_path[0:5] + '_utt'
    worksheet['E1'] = file_path[0:5] + '_annot_utt'

    # Adding english data to the sheet
    for index, row in df_pivot.iterrows():
        # Column 1
        cell = worksheet.cell(row=index + 2, column=1)
        cell.value = row['id'] + 1
        # Column 2
        cell = worksheet.cell(row=index + 2, column=2)
        cell.value = row['utt']
        # Column 3
        cell = worksheet.cell(row=index + 2, column=3)
        cell.value = row['annot_utt']

    # Adding the specified language's data to the sheet
    for index, row in df_lang.iterrows():
        # Column 2
        cell = worksheet.cell(row=index + 2, column=4)
        cell.value = row['utt']
        # Column 3
        cell = worksheet.cell(row=index + 2, column=5)
        cell.value = row['annot_utt']

    # Export the workbook to the specified location
    workbook.save('outputs/xlsx/en-' + language_abbv + '.xlsx')


def filtered_languages():
    # This function selects three languages: English, Swahili and German, and filters them based on the partition
    # column namely test, train and dev respectively

    # Import English, Swahili and German into dataframes
    df_english = pd.read_json(path_or_buf='dataset/data/en-US.jsonl', lines=True)
    df_swahili = pd.read_json(path_or_buf='dataset/data/sw-KE.jsonl', lines=True)
    df_german = pd.read_json(path_or_buf='dataset/data/de-DE.jsonl', lines=True)

    # Next, we filter the imported data to match our requirements
    df_english = df_english[df_english['partition'] == 'test']
    df_swahili = df_swahili[df_swahili['partition'] == 'train']
    df_german = df_german[df_german['partition'] == 'dev']

    # Now, we have to update the indexing on the `id` column before we export the files

    # Loop through all the rows and get their respective translations
    # Variable `id` is used to update the id variable of the new filtered dataframe
    id = 1
    for index, row in df_english.iterrows():
        df_english.at[index, 'id'] = id
        id += 1

    id = 1
    for index, row in df_swahili.iterrows():
        df_swahili.at[index, 'id'] = id

        id += 1

    id = 1
    for index, row in df_german.iterrows():
        df_german.at[index, 'id'] = id

        id += 1
    # From the dataframes, filter and export them into JSONL files.
    df_english.to_json(r'outputs/en-US-test.jsonl', orient='records', lines=True)
    df_swahili.to_json(r'outputs/sw-KE-train.jsonl', orient='records', lines=True)
    df_german.to_json(r'outputs/de-DE-dev.jsonl', orient='records', lines=True)




def get_translations(index):
    # This function returns a dictionary of all translations of the `utt` column based on the index given
    translations_list = {}

    # iterate over all JSONL files in the directory above
    for dataframe in variables:
        # English is skipped since it is a pivot language
        if dataframe['name'].__contains__('en-US'):
            continue

        # Abbv is our key for our dictionary
        abbv = dataframe['name'][0:5]

        # translation is our value pair for the dictionary
        translation: str = dataframe['data'][dataframe['data']['partition'] == 'train'].at[index, 'utt']

        # Now, we add both the key and value pair to the dictionary
        translations_list[abbv] = translation

    return translations_list


def english_translations_to_json():
    # This function creates one large JSONL file with English as the Pivot language and provides translations
    # to all languages in our dataset

    # Set our pivot language, English
    df_pivot: DataFrame = pd.read_json(path_or_buf='dataset/data/en-US.jsonl', lines=True, encoding='utf-8')

    # We filter our pivot language to only retain two rows and to only have the train sets
    pivot_filtered: DataFrame = df_pivot[df_pivot['partition'] == 'train'].get(['id', 'utt'])

    # Create a new empty column called translations which will be an array of all the different translations
    pivot_filtered['translations'] = 'null'

    # Variable `id` is used to update the id variable of the new filtered dataframe
    id = 1
    # Loop through all the rows and get their respective translations
    for index, row in pivot_filtered.iterrows():
        # if index > 1:
        #     break
        pivot_filtered.at[index, 'translations'] = get_translations(index)
        print(id)
        pivot_filtered.at[index, 'id'] = id
        id += 1

    # Export the dataframe to JSONL
    pivot_filtered.to_json(r'outputs/en-xx-train-set-translations.jsonl', orient='records', lines=True, indent=4,
                           force_ascii=False)

